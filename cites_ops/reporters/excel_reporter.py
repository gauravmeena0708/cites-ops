from pathlib import Path
from typing import Any, Dict, Optional, Union
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    """
    Generates professional, multi-tab Excel workbooks (.xlsx) with
    formatted headers, KPI cards, category pivots, and full issue datasets.
    """

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True)
    REGULAR_FONT = Font(name="Calibri", size=11)
    
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    @classmethod
    def generate_report(
        cls,
        df_classified: pd.DataFrame,
        output_path: Union[str, Path],
        workload_data: Optional[Dict[str, Any]] = None,
        title: str = "CITES Operations Intelligence Report",
    ) -> str:
        out_file = Path(output_path)
        out_file.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # 1. Summary Sheet
        ws_summary = wb.create_sheet(title="Executive Summary")
        cls._build_summary_sheet(ws_summary, df_classified, title)

        # 2. Major Categories Sheet
        ws_cat = wb.create_sheet(title="Category Breakdown")
        cls._build_category_sheet(ws_cat, df_classified)

        # 3. All Issues Sheet
        ws_issues = wb.create_sheet(title="All Classified Issues")
        cls._build_issues_sheet(ws_issues, df_classified)

        wb.save(out_file)
        return str(out_file)

    @classmethod
    def _build_summary_sheet(cls, ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame, title: str) -> None:
        ws.views.sheetView[0].showGridLines = True
        ws.merge_cells("A1:E1")
        ws["A1"] = title
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        ws["A1"].alignment = Alignment(vertical="center")

        total = len(df)
        status_series = df["Status"].astype(str).str.lower() if "Status" in df.columns else pd.Series([])
        resolved = int(status_series.isin(["resolved", "fixed", "closed"]).sum())
        open_cnt = total - resolved

        kpis = [
            ("Total Issues Ingested", total),
            ("Open Issues", open_cnt),
            ("Resolved / Closed Issues", resolved),
            ("Resolution Rate", f"{round(100 * resolved / total, 1)}%" if total else "0%"),
        ]

        ws.cell(row=3, column=1, value="Key Performance Indicator").font = cls.BOLD_FONT
        ws.cell(row=3, column=2, value="Count / Value").font = cls.BOLD_FONT
        ws.cell(row=3, column=1).fill = cls.SUBHEADER_FILL
        ws.cell(row=3, column=2).fill = cls.SUBHEADER_FILL

        for idx, (label, val) in enumerate(kpis, start=4):
            c1 = ws.cell(row=idx, column=1, value=label)
            c2 = ws.cell(row=idx, column=2, value=val)
            c1.font = cls.REGULAR_FONT
            c2.font = cls.BOLD_FONT
            c1.border = cls.THIN_BORDER
            c2.border = cls.THIN_BORDER

        # Major Categories Summary
        if "major_topic_label" in df.columns:
            ws.cell(row=10, column=1, value="Major Category Distribution").font = cls.BOLD_FONT
            cat_counts = df["major_topic_label"].value_counts().reset_index()
            cat_counts.columns = ["Major Category", "Count"]

            ws.cell(row=11, column=1, value="Major Category").font = cls.BOLD_FONT
            ws.cell(row=11, column=2, value="Issue Count").font = cls.BOLD_FONT
            ws.cell(row=11, column=1).fill = cls.SUBHEADER_FILL
            ws.cell(row=11, column=2).fill = cls.SUBHEADER_FILL

            for idx, row in cat_counts.iterrows():
                r = 12 + idx
                c1 = ws.cell(row=r, column=1, value=row["Major Category"])
                c2 = ws.cell(row=r, column=2, value=int(row["Count"]))
                c1.font = cls.REGULAR_FONT
                c2.font = cls.REGULAR_FONT
                c1.border = cls.THIN_BORDER
                c2.border = cls.THIN_BORDER

        cls._auto_fit_columns(ws)

    @classmethod
    def _build_category_sheet(cls, ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame) -> None:
        ws.views.sheetView[0].showGridLines = True
        cols = ["major_topic_label", "topic_label", "rule_id"]
        avail_cols = [c for c in cols if c in df.columns]

        if not avail_cols:
            return

        grouped = df.groupby(avail_cols).size().reset_index(name="Issue Count")
        grouped.sort_values(by="Issue Count", ascending=False, inplace=True)

        headers = ["Major Category", "Minor Topic", "Rule Code", "Total Issues"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(grouped.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cls.REGULAR_FONT
                cell.border = cls.THIN_BORDER

        cls._auto_fit_columns(ws)

    @classmethod
    def _build_issues_sheet(cls, ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame) -> None:
        ws.views.sheetView[0].showGridLines = True
        # Select key columns to display cleanly
        pref_cols = [
            "Id", "Category", "Status", "Assigned To", "Date Submitted",
            "major_topic_label", "topic_label", "workflow_level_label", "rule_id", "Summary"
        ]
        export_cols = [c for c in pref_cols if c in df.columns] + [c for c in df.columns if c not in pref_cols and not c.startswith("_")]

        # Write Headers
        for col_idx, col_name in enumerate(export_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write Data
        for row_idx, row in enumerate(df[export_cols].itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                cell.font = cls.REGULAR_FONT
                cell.border = cls.THIN_BORDER

        cls._auto_fit_columns(ws)

    @classmethod
    def _auto_fit_columns(cls, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    # Limit line length evaluation for long text
                    first_line = val_str.split("\n")[0]
                    max_len = max(max_len, min(len(first_line), 60))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
